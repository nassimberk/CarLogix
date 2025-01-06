import flet as ft
from flet import icons, colors
import openpyxl
from openpyxl import Workbook
from datetime import datetime, timedelta
import re
import os
import pandas as pd
from dataclasses import dataclass
from typing import List, Optional
import tkinter as tk
from tkinter import filedialog
from reportlab.lib import colors as pdf_colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.units import inch
import subprocess
from pydantic import BaseModel, ValidationError, validator

@dataclass
class Vehicle:
    id: int
    immatriculation: str
    code_carte: str
    societe_proprietaire: str
    site: str
    utilisateur: str
    marque: str
    vehicule: str
    modele: str
    date_mise_en_service: str
    crit_air: str
    carburant: str
    type_huile: str
    fluide_dispo: str
    releve_kms: str
    date_derniere_revision: str
    derniere_revision: str
    periodicite_revision: str
    prochain_ct: str
    double_clef: str
    numero_scelle: str
    statut: str

class VehicleModel(BaseModel):
    id: Optional[int] = None
    immatriculation: str
    code_carte: str
    societe_proprietaire: str
    site: str
    utilisateur: str
    marque: str
    vehicule: str
    modele: str
    date_mise_en_service: str
    crit_air: str
    carburant: str
    type_huile: str
    fluide_dispo: str
    releve_kms: str
    date_derniere_revision: str
    derniere_revision: str
    periodicite_revision: str
    prochain_ct: str
    double_clef: str
    numero_scelle: str
    statut: str

    ('immatriculation')
    def validate_immatriculation(cls, value):
        if not re.match(r'^[A-Za-z]{2}-\d{3}-[A-Za-z]{2}$', value):
            raise ValueError('L\'immatriculation doit être sous format AB-123-CD')
        return value

    ('code_carte')
    def validate_code_carte(cls, value):
        if not re.match(r'^\d{4}$', value):
            raise ValueError('Le code carte doit contenir 4 chiffres')
        return value

    ('releve_kms', 'derniere_revision', 'numero_scelle')
    def validate_numeric(cls, value):
        if not re.match(r'^\d+$', value):
            raise ValueError('Ce champ doit contenir uniquement des chiffres')
        return value

class VehicleRepository:
    def __init__(self, file_path="CarLogix_DATA.xlsx"):
        self.file_path = file_path
        if not os.path.exists(self.file_path):
            self.create_excel_file()

    def create_excel_file(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Vehicules"
        ws.append([
            "ID", "Immatriculation", "Code Carte", "Société Propriétaire", "Site", "Utilisateur",
            "Marque", "Véhicule", "Modèle", "Date de mise en service", "CRIT AIR", "Carburant",
            "Type Huile", "Fluide Dispo", "Relevé KMS", "Date Dernière Révision",
            "Dernière Révision", "Périodicité Révision", "Prochain C.T.", "Double de clef", "N° Scellé du double", "Statut",
        ])
        wb.save(self.file_path)

    def fetch_all_vehicles(self) -> List[Vehicle]:
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb.active
        vehicles = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            vehicles.append(Vehicle(*row))
        return vehicles

    def add_vehicle(self, vehicle: VehicleModel) -> int:
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb.active
        last_id = ws.max_row - 1
        new_id = last_id + 1
        ws.append([
            new_id, vehicle.immatriculation, vehicle.code_carte, vehicle.societe_proprietaire, vehicle.site,
            vehicle.utilisateur, vehicle.marque, vehicle.vehicule, vehicle.modele, vehicle.date_mise_en_service,
            vehicle.crit_air, vehicle.carburant, vehicle.type_huile, vehicle.fluide_dispo, vehicle.releve_kms,
            vehicle.date_derniere_revision, vehicle.derniere_revision, vehicle.periodicite_revision,
            vehicle.prochain_ct, vehicle.double_clef, vehicle.numero_scelle, vehicle.statut
        ])
        wb.save(self.file_path)
        return new_id

    def update_vehicle(self, vehicle: VehicleModel):
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if row[0].value == vehicle.id:
                row[1].value = vehicle.immatriculation
                row[2].value = vehicle.code_carte
                row[3].value = vehicle.societe_proprietaire
                row[4].value = vehicle.site
                row[5].value = vehicle.utilisateur
                row[6].value = vehicle.marque
                row[7].value = vehicle.vehicule
                row[8].value = vehicle.modele
                row[9].value = vehicle.date_mise_en_service
                row[10].value = vehicle.crit_air
                row[11].value = vehicle.carburant
                row[12].value = vehicle.type_huile
                row[13].value = vehicle.fluide_dispo
                row[14].value = vehicle.releve_kms
                row[15].value = vehicle.date_derniere_revision
                row[16].value = vehicle.derniere_revision
                row[17].value = vehicle.periodicite_revision
                row[18].value = vehicle.prochain_ct
                row[19].value = vehicle.double_clef
                row[20].value = vehicle.numero_scelle
                row[21].value = vehicle.statut
                break
        wb.save(self.file_path)

    def delete_vehicle(self, id: int):
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if row[0].value == id:
                ws.delete_rows(row[0].row)
                break
        wb.save(self.file_path)

class VehicleManagementApp:
    FIELD_WIDTH = 200
    FIELD_HEIGHT = 45

    CRIT_AIR_OPTIONS = ["0", "1", "2", "3", "4", "5"]
    CARBURANT_OPTIONS = ["Essence", "Diesel", "Électrique", "Hybride", "GPL"]
    STATUT_OPTIONS = ["En service", "En maintenance", "Hors service"]
    SITE_OPTIONS = ["ELAN", "DSO", "DSE", "DNE"]
    FLUIDE_DISPO_OPTIONS = ["Oui", "Non"]
    DOUBLE_CLE_OPTIONS = ["Oui", "Non"]
    SOCIETE_PROPRIETAIRE_OPTIONS = ["JIVAGO", "ARVAL"]

    def __init__(self, page: ft.Page, vehicle_repository: VehicleRepository):
        self.page = page
        self.page.title = "CarLogix"
        self.page.icon = "assets/icon.png"
        self.page.theme_mode = ft.ThemeMode.LIGHT
        self.page.padding = 0
        self.vehicle_repository = vehicle_repository
        self.setup_page()
        self.error_style = ft.TextStyle(color="red")

    def create_date_picker(self, label: str, hint_text: str = "JJ/MM/AAAA", value: Optional[str] = None) -> ft.Container:
        date_picker = ft.DatePicker(
            first_date=datetime(2000, 1, 1),
            last_date=datetime(2080, 12, 31),
        )

        text_field = ft.TextField(
            label=label,
            value=value,
            read_only=True,
            width=self.FIELD_WIDTH,
            height=self.FIELD_HEIGHT,
            hint_text=hint_text,
        )

        def date_changed(e):
            if date_picker.value:
                text_field.value = date_picker.value.strftime('%d/%m/%Y')
                text_field.update()

        date_picker.on_change = date_changed

        icon_button = ft.IconButton(
            icon=ft.icons.CALENDAR_TODAY,
            icon_size=20,
            on_click=lambda _: date_picker.pick_date()
        )

        container = ft.Container(
            content=ft.Row(
                controls=[text_field, icon_button],
                spacing=10,
                alignment=ft.MainAxisAlignment.START,
            ),
            data={"text_field": text_field, "date_picker": date_picker},
        )

        self.page.controls.append(date_picker)
        self.page.update()

        return container

    def create_code_carte_field(self, width: int, initial_value: Optional[str] = None) -> ft.TextField:
        return ft.TextField(
            label="Code Carte",
            value=str(initial_value)[:4] if initial_value else None,
            width=width,
            height=self.FIELD_HEIGHT,
            keyboard_type="number",
            hint_text="Entrez 4 chiffres",
            text_align=ft.TextAlign.CENTER,
        )

    def create_required_field(self, label: str, width: int = FIELD_WIDTH) -> ft.TextField:
        return ft.TextField(
            label=f"{label} *",
            width=width,
            label_style=ft.TextStyle(color="red"),
            border_color="red"
        )

    def create_dropdown(self, label: str, options: List[str], value: Optional[str] = None,
                        width: int = FIELD_WIDTH, height: int = FIELD_HEIGHT) -> ft.Dropdown:
        return ft.Dropdown(
            label=label,
            width=width,
            height=height,
            value=value,
            options=[ft.dropdown.Option(opt) for opt in options]
        )

    def validate_numeric(self, e):
        current_value = e.control.value
        cleaned_value = re.sub(r'[^0-9]', '', current_value)

        if current_value != cleaned_value:
            e.control.value = cleaned_value
            e.control.update()

    def setup_page(self):
        self.search_field = ft.TextField(
            hint_text="Rechercher par Immatriculation, Marque, Modèle ou Utilisateur...",
            on_change=self.search_vehicles,
            width=400,
            suffix_icon=ft.icons.SEARCH
        )

        self.page.appbar = ft.AppBar(
            leading=ft.Icon(icons.DIRECTIONS_CAR),
            leading_width=40,
            title=ft.Row(
                [
                    self.search_field
                ],
                alignment=ft.MainAxisAlignment.CENTER,
                expand=True
            ),
            center_title=False,
            bgcolor=ft.colors.SURFACE_VARIANT,
            actions=[
                ft.ElevatedButton(text="Nouveau véhicule", bgcolor="#424242", color="#f7f9f9", on_click=self.add_vehicle),
                ft.IconButton(ft.icons.SETTINGS, icon_size=35, on_click=self.show_settings)
            ],
        )

        self.navigation_rail = ft.NavigationRail(
            selected_index=0,
            label_type=ft.NavigationRailLabelType.ALL,
            min_width=100,
            min_extended_width=400,
            destinations=[
                ft.NavigationRailDestination(
                    icon=icons.HOME, label="Accueil"
                ),
                ft.NavigationRailDestination(
                    icon=icons.DIRECTIONS_CAR, label="Véhicules"
                ),
                ft.NavigationRailDestination(
                    icon=icons.BUILD, label="Entretien"
                ),
                ft.NavigationRailDestination(
                    icon=icons.FACT_CHECK, label="Contrôles Techniques"
                ),
            ],
            on_change=self.change_tab
        )

        self.stats_view = ft.Container(
            content=self.create_stats_view(),
            padding=20,
        )

        self.vehicles_view = ft.ListView(spacing=10, padding=20, auto_scroll=True)
        self.update_vehicles_list()

        self.main_content = ft.Container(
            content=self.stats_view,
            expand=True,
        )

        self.page.add(
            ft.Row(
                [
                    self.navigation_rail,
                    ft.VerticalDivider(width=1),
                    self.main_content,
                ],
                expand=True,
            )
        )

    def create_stats_view(self):
        vehicles = self.vehicle_repository.fetch_all_vehicles()
        if not vehicles:
            return ft.Text("Aucun véhicule disponible.", size=20, color=colors.RED, weight=ft.FontWeight.BOLD)

        total_kms = sum(int(vehicle.releve_kms) for vehicle in vehicles if isinstance(vehicle.releve_kms, (int, float)))
        total_vehicles = len(vehicles)
        avg_kms = total_kms / total_vehicles if total_vehicles > 0 else 0

        avg_monthly_kms = avg_kms / 12
        avg_co2_emission = 110  # Placeholder value
        avg_age = 15  # Placeholder value

        vehicle_data = [vars(vehicle) for vehicle in vehicles]
        if vehicle_data:
            marque_distribution = pd.DataFrame(vehicle_data)["marque"].value_counts(normalize=True) * 100
            carburant_distribution = pd.DataFrame(vehicle_data)["carburant"].value_counts(normalize=True) * 100
            societe_distribution = pd.DataFrame(vehicle_data)["societe_proprietaire"].value_counts(normalize=True) * 100
            site_distribution = pd.DataFrame(vehicle_data)["site"].value_counts()
        else:
            marque_distribution = pd.Series(dtype=float)
            carburant_distribution = pd.Series(dtype=float)
            societe_distribution = pd.Series(dtype=float)
            site_distribution = pd.Series(dtype=float)

        maintenance_count = self.calculate_maintenance_count()
        ct_count = self.calculate_ct_count()

        return ft.Column([
            ft.Row([
                ft.Card(
                    content=ft.Container(
                        content=ft.Column([
                            ft.Text("Nombre de véhicules", size=18, weight=ft.FontWeight.BOLD),
                            ft.Text(f"{total_vehicles} véhicules", size=15),
                        ]),
                        padding=20,
                    ),
                    width=200,
                    height=220,
                ),
                ft.Card(
                    content=ft.Container(
                        content=ft.Column([
                            ft.Text("Répartition par marque", size=18, weight=ft.FontWeight.BOLD),
                            ft.Text("\n".join([f"{marque}: {percent:.2f}%" for marque, percent in marque_distribution.items()]), size=15),
                        ]),
                        padding=20,
                    ),
                    width=200,
                    height=220,
                ),
                ft.Card(
                    content=ft.Container(
                        content=ft.Column([
                            ft.Text("Répartition par carburant", size=18, weight=ft.FontWeight.BOLD),
                            ft.Text("\n".join([f"{carburant}: {percent:.2f}%" for carburant, percent in carburant_distribution.items()]), size=14),
                        ]),
                        padding=20,
                    ),
                    width=200,
                    height=220,
                ),
                ft.Card(
                    content=ft.Container(
                        content=ft.Column([
                            ft.Text("Répartition par société", size=18, weight=ft.FontWeight.BOLD),
                            ft.Text("\n".join([f"{societe_distribution}: {percent:.2f}%" for societe_distribution, percent in societe_distribution.items()]), size=14),
                        ]),
                        padding=20,
                    ),
                    width=200,
                    height=220,
                ),
                ft.Card(
                    content=ft.Container(
                        content=ft.Column([
                            ft.Text("Répartition par site", size=18, weight=ft.FontWeight.BOLD),
                            ft.Text("\n".join([f"{site}: {count}" for site, count in site_distribution.items()]),size=12),
                        ]),
                        padding=20,
                    ),
                    width=200,
                    height=220,
                ),
                ft.Card(
                    content=ft.Container(
                        content=ft.Column([
                            ft.Text("Véhicules nécessitant un entretien", size=18, weight=ft.FontWeight.BOLD),
                            ft.Text(f"{maintenance_count} véhicules", size=15),
                        ]),
                        padding=20,
                    ),
                    width=200,
                    height=220,
                ),
                ft.Card(
                    content=ft.Container(
                        content=ft.Column([
                            ft.Text("Véhicules nécessitant un contrôle technique", size=18, weight=ft.FontWeight.BOLD),
                            ft.Text(f"{ct_count} véhicules", size=15),
                        ]),
                        padding=20,
                    ),
                    width=200,
                    height=220,
                ),
            ], spacing=20),
            ft.Row([
                ft.Card(
                    content=ft.Container(
                        content=ft.Column([
                            ft.Text("Véhicules nécessitant un contrôle technique", size=18, weight=ft.FontWeight.BOLD),
                            ft.Text(f"{ct_count} véhicules", size=15),
                        ]),
                        padding=20,
                    ),
                    width=200,
                    height=220,
                ),
                ft.Card(
                    content=ft.Container(
                        content=ft.Column([
                            ft.Text("Véhicules nécessitant un contrôle technique", size=18, weight=ft.FontWeight.BOLD),
                            ft.Text(f"{ct_count} véhicules", size=15),
                        ]),
                        padding=20,
                    ),
                    width=200,
                    height=220,
                ),

            ], spacing=20),
        ], spacing=20)

    def update_vehicles_list(self, search_text: str = ""):
        vehicles = self.vehicle_repository.fetch_all_vehicles()
        status_colors = {
            "En service": colors.GREEN,
            "En maintenance": colors.ORANGE,
            "Hors service": colors.RED,
            "En attente": colors.BLUE,
        }

        self.vehicles_view.controls.clear()

        for vehicle in vehicles:
            if search_text.lower() in f"{vehicle.immatriculation} {vehicle.marque} {vehicle.vehicule} {vehicle.utilisateur} {vehicle.site}".lower():
                status_color = status_colors.get(vehicle.statut, colors.GREY)

                vehicle_card = ft.Card(
                    content=ft.Container(
                        content=ft.Column([
                            ft.ListTile(
                                leading=ft.Icon(icons.DIRECTIONS_CAR, size=40, color=colors.BLUE),
                                height=60,
                                title=ft.Text(
                                    f"{vehicle.marque} {vehicle.vehicule} // {vehicle.immatriculation} ",
                                    size=20,
                                    weight=ft.FontWeight.BOLD
                                ),
                                subtitle=ft.Column([
                                    ft.Text(f"Utilisateur: {vehicle.utilisateur}", size=10, weight=ft.FontWeight.BOLD),
                                    ft.Text(f"Site: {vehicle.site}", size=10, weight=ft.FontWeight.BOLD),
                                    ft.Container(
                                        content=ft.Text(
                                            vehicle.statut,
                                            color=colors.WHITE,
                                            size=8,
                                            weight=ft.FontWeight.BOLD
                                        ),
                                        bgcolor=status_color,
                                        padding=10,
                                        border_radius=15,
                                    ),
                                ]),
                            ),
                            ft.Row(
                                [
                                    ft.TextButton(
                                        "Modifier",
                                        on_click=lambda _, id=vehicle.id: self.edit_vehicle(id)
                                    ),
                                    ft.TextButton(
                                        "Supprimer",
                                        on_click=lambda _, id=vehicle.id: self.delete_vehicle(id)
                                    ),
                                    ft.TextButton(
                                        "Détails",
                                        icon=ft.icons.INFO,
                                        on_click=lambda _, id=vehicle.id: self.show_vehicle_details(id)
                                    ),
                                ],
                                alignment=ft.MainAxisAlignment.END,
                            ),
                        ]),
                        padding=5,
                    )
                )

                self.vehicles_view.controls.append(vehicle_card)

        self.page.update()

    def show_vehicle_details(self, vehicle_id: int):
        vehicle = next((v for v in self.vehicle_repository.fetch_all_vehicles() if v.id == vehicle_id), None)
        if not vehicle:
            return

        details_dialog = ft.AlertDialog(
            title=ft.Text(f"Détails du véhicule - {vehicle.immatriculation}"),
            content=ft.Column([
                self.create_detail_section("Informations générales", [
                    ("Marque", vehicle.marque),
                    ("Véhicule", vehicle.vehicule),
                    ("Mise en service", vehicle.date_mise_en_service),
                    ("Utilisateur", vehicle.utilisateur),
                    ("Site", vehicle.site),
                    ("Société propriétaire", vehicle.societe_proprietaire),
                    ("Code CC", vehicle.code_carte),
                ]),
                self.create_detail_section("Technique", [
                    ("Modèle", vehicle.modele),
                    ("CRIT AIR", vehicle.crit_air),
                    ("Carburant", vehicle.carburant),
                    ("Type Huile", vehicle.type_huile),
                    ("Fluide Dispo", vehicle.fluide_dispo),
                ]),
                self.create_detail_section("Entretien", [
                    ("Statut", vehicle.statut),
                    ("Dernier relevé KMS", vehicle.releve_kms),
                    ("Dernière révision", vehicle.date_derniere_revision),
                    ("Dernière révision (kms)", vehicle.derniere_revision),
                    ("Périodicité révision", vehicle.periodicite_revision),
                ]),
                self.create_detail_section("Sécurité", [
                    ("Prochain CT", vehicle.prochain_ct),
                    ("Double de clef", vehicle.double_clef),
                    ("N° Scellé", vehicle.numero_scelle),
                ]),
            ], spacing=10, scroll=ft.ScrollMode.AUTO),
            actions=[
                ft.TextButton("Fermer", on_click=lambda _: setattr(details_dialog, 'open', False))
            ],
        )

        self.page.dialog = details_dialog
        details_dialog.open = True
        self.page.update()

    def create_detail_section(self, title: str, items: List[tuple]) -> ft.Container:
        return ft.Container(
            content=ft.Column([
                ft.Text(title, size=18, weight=ft.FontWeight.BOLD),
                ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Text(label + ":", weight=ft.FontWeight.BOLD),
                            ft.Text(str(value))
                        ])
                        for label, value in items
                    ], spacing=5),
                    padding=10
                )
            ]),
            bgcolor=colors.SURFACE_VARIANT,
            border_radius=10,
            padding=10,
        )

    def search_vehicles(self, e):
        search_text = self.search_field.value
        self.update_vehicles_list(search_text)

    def add_vehicle(self, e):
        def save_vehicle(e):
            try:
                vehicle = VehicleModel(
                    immatriculation=immatriculation_field.value,
                    code_carte=code_carte_field.value,
                    societe_proprietaire=societe_proprietaire_field.value,
                    site=site_field.value,
                    utilisateur=utilisateur_field.value,
                    marque=marque_field.value,
                    vehicule=vehicule_field.value,
                    modele=modele_field.value,
                    date_mise_en_service=date_mise_en_service_field.data["text_field"].value,
                    crit_air=crit_air_field.value,
                    carburant=carburant_field.value,
                    type_huile=type_huile_field.value,
                    fluide_dispo=fluide_dispo_field.value,
                    releve_kms=releve_kms_field.value,
                    date_derniere_revision=date_derniere_revision_field.data["text_field"].value,
                    derniere_revision=derniere_revision_field.value,
                    periodicite_revision=periodicite_revision_field.value,
                    prochain_ct=prochain_ct_field.data["text_field"].value,
                    double_clef=double_clef_field.value,
                    numero_scelle=numero_scelle_field.value,
                    statut=statut_field.value,
                )
                self.vehicle_repository.add_vehicle(vehicle)
                dialog.open = False
                self.page.update()
                self.update_vehicles_list()
                self.create_stats_view()
                self.page.update()
            except ValidationError as e:
                self._show_error_dialog(str(e))

        immatriculation_field = ft.TextField(
            label="Immatriculation",
            width=self.FIELD_WIDTH,
            height=self.FIELD_HEIGHT,
            hint_text="AB-123-CD"
        )
        code_carte_field = self.create_code_carte_field(self.FIELD_WIDTH)
        societe_proprietaire_field = self.create_dropdown("Société Propriétaire", self.SOCIETE_PROPRIETAIRE_OPTIONS)
        site_field = self.create_dropdown("SITE", self.SITE_OPTIONS)
        utilisateur_field = ft.TextField(label="Utilisateur", width=self.FIELD_WIDTH, height=self.FIELD_HEIGHT)
        marque_field = ft.TextField(label="Marque", width=self.FIELD_WIDTH, height=self.FIELD_HEIGHT)
        vehicule_field = ft.TextField(label="Véhicule", width=self.FIELD_WIDTH, height=self.FIELD_HEIGHT)
        modele_field = ft.TextField(label="Modèle", width=self.FIELD_WIDTH, height=self.FIELD_HEIGHT)
        date_mise_en_service_field = self.create_date_picker("Date mise en service", hint_text="JJ/MM/AAAA")
        crit_air_field = self.create_dropdown("CRIT AIR", self.CRIT_AIR_OPTIONS)
        carburant_field = self.create_dropdown("Carburant", self.CARBURANT_OPTIONS)
        type_huile_field = ft.TextField(label="Type Huile", width=self.FIELD_WIDTH, height=self.FIELD_HEIGHT)
        fluide_dispo_field = self.create_dropdown("Fluide dispo", self.FLUIDE_DISPO_OPTIONS)
        releve_kms_field = ft.TextField(
            label="Relevé KMS",
            width=self.FIELD_WIDTH,
            height=self.FIELD_HEIGHT,
            keyboard_type="number",
            hint_text="Entrez que les chiffres",
            on_change=self.validate_numeric,
            error_text=""
        )
        date_derniere_revision_field = self.create_date_picker("Date Dernière Révision", hint_text="JJ/MM/AA")
        derniere_revision_field = ft.TextField(
            label="Dernière Révision KMS",
            width=self.FIELD_WIDTH,
            height=self.FIELD_HEIGHT,
            keyboard_type="number",
            hint_text="Entrez que les chiffres",
            on_change=self.validate_numeric,
            error_text=""
        )
        periodicite_revision_field = ft.TextField(label="Périodicité Révision", width=self.FIELD_WIDTH, height=self.FIELD_HEIGHT)
        prochain_ct_field = self.create_date_picker("Prochain C.T", hint_text="JJ/MM/AA")
        double_clef_field = self.create_dropdown("Double de clef", self.DOUBLE_CLE_OPTIONS)
        numero_scelle_field = ft.TextField(
            label="N° Scellé du double",
            keyboard_type="number",
            hint_text="Entrez que les chiffres",
            width=self.FIELD_WIDTH,
            height=self.FIELD_HEIGHT,
            on_change=self.validate_numeric,
            error_text=""
        )
        statut_field = self.create_dropdown("Statut", self.STATUT_OPTIONS)

        dialog_content = ft.Column([
            ft.Row([immatriculation_field, code_carte_field, societe_proprietaire_field], alignment=ft.MainAxisAlignment.CENTER),
            ft.Row([site_field, utilisateur_field, marque_field], alignment=ft.MainAxisAlignment.CENTER),
            ft.Row([vehicule_field, modele_field, date_mise_en_service_field], alignment=ft.MainAxisAlignment.CENTER),
            ft.Row([crit_air_field, carburant_field, type_huile_field], alignment=ft.MainAxisAlignment.CENTER),
            ft.Row([fluide_dispo_field, releve_kms_field, date_derniere_revision_field], alignment=ft.MainAxisAlignment.CENTER),
            ft.Row([derniere_revision_field, periodicite_revision_field, prochain_ct_field], alignment=ft.MainAxisAlignment.CENTER),
            ft.Row([double_clef_field, numero_scelle_field, statut_field], alignment=ft.MainAxisAlignment.CENTER)
        ], alignment=ft.MainAxisAlignment.CENTER, spacing=10)

        dialog = ft.AlertDialog(
            title=ft.Text("Ajouter un véhicule"),
            content=ft.Container(
                content=dialog_content,
                padding=20,
                width=700
            ),
            actions=[
                ft.TextButton("Annuler", on_click=lambda _: setattr(dialog, 'open', False)),
                ft.TextButton("Sauvegarder", on_click=save_vehicle),
            ],
        )

        self.page.dialog = dialog
        dialog.open = True
        self.page.update()

    def edit_vehicle(self, vehicle_id: int):
        vehicle = next((v for v in self.vehicle_repository.fetch_all_vehicles() if v.id == vehicle_id), None)
        if not vehicle:
            return

        def save_changes(e):
            try:
                updated_vehicle = VehicleModel(
                    id=vehicle.id,
                    immatriculation=immatriculation_field.value,
                    code_carte=code_carte_field.value,
                    societe_proprietaire=societe_proprietaire_field.value,
                    site=site_field.value,
                    utilisateur=utilisateur_field.value,
                    marque=marque_field.value,
                    vehicule=vehicule_field.value,
                    modele=modele_field.value,
                    date_mise_en_service=date_mise_en_service_field.data["text_field"].value,
                    crit_air=crit_air_field.value,
                    carburant=carburant_field.value,
                    type_huile=type_huile_field.value,
                    fluide_dispo=fluide_dispo_field.value,
                    releve_kms=releve_kms_field.value,
                    date_derniere_revision=date_derniere_revision_field.data["text_field"].value,
                    derniere_revision=derniere_revision_field.value,
                    periodicite_revision=periodicite_revision_field.value,
                    prochain_ct=prochain_ct_field.data["text_field"].value,
                    double_clef=double_clef_field.value,
                    numero_scelle=numero_scelle_field.value,
                    statut=statut_field.value,
                )
                self.vehicle_repository.update_vehicle(updated_vehicle)
                dialog.open = False
                self.page.update()
                self.update_vehicles_list()
                self.create_stats_view()
                self.page.update()
            except ValidationError as e:
                self._show_error_dialog(str(e))

        immatriculation_field = ft.TextField(
            label="Immatriculation",
            value=vehicle.immatriculation,
            width=self.FIELD_WIDTH,
            height=self.FIELD_HEIGHT,
            hint_text="AB-123-CD"
        )
        code_carte_field = self.create_code_carte_field(self.FIELD_WIDTH, initial_value=vehicle.code_carte)
        societe_proprietaire_field = self.create_dropdown("Société Propriétaire", self.SOCIETE_PROPRIETAIRE_OPTIONS, value=vehicle.societe_proprietaire)
        site_field = self.create_dropdown("SITE", self.SITE_OPTIONS, value=vehicle.site)
        utilisateur_field = ft.TextField(label="Utilisateur", value=vehicle.utilisateur, width=self.FIELD_WIDTH, height=self.FIELD_HEIGHT)
        marque_field = ft.TextField(label="Marque", value=vehicle.marque, width=self.FIELD_WIDTH, height=self.FIELD_HEIGHT)
        vehicule_field = ft.TextField(label="Véhicule", value=vehicle.vehicule, width=self.FIELD_WIDTH, height=self.FIELD_HEIGHT)
        modele_field = ft.TextField(label="Modèle", value=vehicle.modele, width=self.FIELD_WIDTH, height=self.FIELD_HEIGHT)
        date_mise_en_service_field = self.create_date_picker("Date de mise en service", hint_text="JJ/MM/AA", value=vehicle.date_mise_en_service)
        crit_air_field = self.create_dropdown("CRIT AIR", self.CRIT_AIR_OPTIONS, value=vehicle.crit_air)
        carburant_field = self.create_dropdown("Carburant", self.CARBURANT_OPTIONS, value=vehicle.carburant)
        type_huile_field = ft.TextField(label="Type Huile", value=vehicle.type_huile, width=self.FIELD_WIDTH, height=self.FIELD_HEIGHT)
        fluide_dispo_field = self.create_dropdown("Fluide dispo", self.FLUIDE_DISPO_OPTIONS, value=vehicle.fluide_dispo)
        date_derniere_revision_field = self.create_date_picker("Date Dernière Révision", hint_text="JJ/MM/AA", value=vehicle.date_derniere_revision)
        releve_kms_field = ft.TextField(
            label="Relevé KMS",
            width=self.FIELD_WIDTH,
            height=self.FIELD_HEIGHT,
            keyboard_type="number",
            hint_text="Entrez que les chiffres",
            value=vehicle.releve_kms,
            on_change=self.validate_numeric,
            error_text=""
        )
        derniere_revision_field = ft.TextField(
            label="Dernière Révision KMS",
            width=self.FIELD_WIDTH,
            height=self.FIELD_HEIGHT,
            keyboard_type="number",
            hint_text="Entrez que les chiffres",
            value=vehicle.derniere_revision,
            on_change=self.validate_numeric,
            error_text=""
        )
        periodicite_revision_field = ft.TextField(label="Périodicité Révision", value=vehicle.periodicite_revision, width=self.FIELD_WIDTH)
        prochain_ct_field = self.create_date_picker(label="Prochain C.T.", hint_text="JJ/MM/AA", value=vehicle.prochain_ct)
        double_clef_field = self.create_dropdown("Double de clef", self.DOUBLE_CLE_OPTIONS, value=vehicle.double_clef)
        numero_scelle_field = ft.TextField(
            label="N° Scellé du double",
            keyboard_type="number",
            hint_text="Entrez que les chiffres",
            width=self.FIELD_WIDTH,
            height=self.FIELD_HEIGHT,
            value=vehicle.numero_scelle,
            on_change=self.validate_numeric,
            error_text=""
        )
        statut_field = self.create_dropdown("Statut", self.STATUT_OPTIONS, value=vehicle.statut)

        dialog_content = ft.Column([
            ft.Row([immatriculation_field, code_carte_field, societe_proprietaire_field], alignment=ft.MainAxisAlignment.CENTER),
            ft.Row([site_field, utilisateur_field, marque_field], alignment=ft.MainAxisAlignment.CENTER),
            ft.Row([vehicule_field, modele_field, date_mise_en_service_field], alignment=ft.MainAxisAlignment.CENTER),
            ft.Row([crit_air_field, carburant_field, type_huile_field], alignment=ft.MainAxisAlignment.CENTER),
            ft.Row([fluide_dispo_field, releve_kms_field, date_derniere_revision_field], alignment=ft.MainAxisAlignment.CENTER),
            ft.Row([derniere_revision_field, periodicite_revision_field, prochain_ct_field], alignment=ft.MainAxisAlignment.CENTER),
            ft.Row([double_clef_field, numero_scelle_field, statut_field], alignment=ft.MainAxisAlignment.CENTER)
        ], alignment=ft.MainAxisAlignment.CENTER, spacing=10)

        dialog = ft.AlertDialog(
            title=ft.Text("Modifier le véhicule"),
            content=ft.Container(
                content=dialog_content,
                padding=20,
                width=700
            ),
            actions=[
                ft.TextButton("Annuler", on_click=lambda _: setattr(dialog, 'open', False)),
                ft.TextButton("Sauvegarder", on_click=save_changes),
            ],
        )

        self.page.dialog = dialog
        dialog.open = True
        self.page.update()

    def delete_vehicle(self, vehicle_id: int):
        def confirm_delete(e):
            self.vehicle_repository.delete_vehicle(vehicle_id)
            confirm_dialog.open = False
            self.update_vehicles_list()
            self.create_stats_view()
            self.page.update()

        confirm_dialog = ft.AlertDialog(
            title=ft.Text("Confirmer la suppression"),
            content=ft.Text("Êtes-vous sûr de vouloir supprimer ce véhicule ?"),
            actions=[
                ft.TextButton("Annuler", on_click=lambda _: setattr(confirm_dialog, 'open', False)),
                ft.TextButton("Supprimer", on_click=confirm_delete),
            ],
        )

        self.page.dialog = confirm_dialog
        confirm_dialog.open = True
        self.page.update()

    def change_tab(self, e):
        if e.control.selected_index == 0:
            self.main_content.content = self.stats_view
        elif e.control.selected_index == 1:
            self.update_vehicles_list()
            self.main_content.content = self.vehicles_view
        elif e.control.selected_index == 2:
            self.show_maintenance_alerts()
        elif e.control.selected_index == 3:
            self.show_ct_alerts()
        self.page.update()

    def show_settings(self, e):
        settings_dialog = ft.AlertDialog(
            title=ft.Text("Paramètres"),
            content=ft.Column([
                ft.ListTile(
                    leading=ft.Icon(icons.DARK_MODE, size=20),
                    title=ft.Text("Mode sombre", size=15),
                    trailing=ft.Switch(
                        value=self.page.theme_mode == ft.ThemeMode.DARK,
                        width=40,
                        height=20,
                        on_change=self.toggle_theme_mode
                    ),
                ),
                ft.ListTile(
                    leading=ft.Icon(icons.BACKUP),
                    title=ft.Text("Exporter les données"),
                    on_click=self.show_export_dialog
                ),
                ft.ListTile(
                    leading=ft.Icon(icons.EDIT),
                    title=ft.Text("Modifier les listes déroulantes"),
                    on_click=self.show_dropdown_edit_dialog
                ),
            ]),
            actions=[
                ft.TextButton("Fermer", on_click=lambda _: setattr(settings_dialog, 'open', False))
            ],
            alignment=ft.alignment.center_right
        )

        self.page.dialog = settings_dialog
        settings_dialog.open = True
        self.page.update()

    def show_dropdown_edit_dialog(self, e):
        dropdown_edit_dialog = ft.AlertDialog(
            title=ft.Text("Modifier les listes déroulantes"),
            content=ft.Column([
                ft.TextButton("Modifier les options de CRIT AIR", on_click=lambda _: self.show_edit_dropdown_dialog("CRIT AIR", self.CRIT_AIR_OPTIONS)),
                ft.TextButton("Modifier les options de Carburant", on_click=lambda _: self.show_edit_dropdown_dialog("Carburant", self.CARBURANT_OPTIONS)),
                ft.TextButton("Modifier les options de Statut", on_click=lambda _: self.show_edit_dropdown_dialog("Statut", self.STATUT_OPTIONS)),
                ft.TextButton("Modifier les options de Site", on_click=lambda _: self.show_edit_dropdown_dialog("Site", self.SITE_OPTIONS)),
                ft.TextButton("Modifier les options de Fluide Dispo", on_click=lambda _: self.show_edit_dropdown_dialog("Fluide Dispo", self.FLUIDE_DISPO_OPTIONS)),
                ft.TextButton("Modifier les options de Double Clef", on_click=lambda _: self.show_edit_dropdown_dialog("Double Clef", self.DOUBLE_CLE_OPTIONS)),
                ft.TextButton("Modifier les options de Société Propriétaire", on_click=lambda _: self.show_edit_dropdown_dialog("Société Propriétaire", self.SOCIETE_PROPRIETAIRE_OPTIONS)),
            ], spacing=10),
            actions=[
                ft.TextButton("Fermer", on_click=lambda _: setattr(dropdown_edit_dialog, 'open', False))
            ],
        )

        self.page.dialog = dropdown_edit_dialog
        dropdown_edit_dialog.open = True
        self.page.update()

    def show_edit_dropdown_dialog(self, dropdown_name: str, options: List[str]):
        options_field = ft.TextField(
            label=f"Options pour {dropdown_name}",
            value="\n".join(options),
            multiline=True,
            height=200,
            width=300,
        )

        def save_options(e):
            new_options = options_field.value.split("\n")
            if dropdown_name == "CRIT AIR":
                self.CRIT_AIR_OPTIONS = new_options
            elif dropdown_name == "Carburant":
                self.CARBURANT_OPTIONS = new_options
            elif dropdown_name == "Statut":
                self.STATUT_OPTIONS = new_options
            elif dropdown_name == "Site":
                self.SITE_OPTIONS = new_options
            elif dropdown_name == "Fluide Dispo":
                self.FLUIDE_DISPO_OPTIONS = new_options
            elif dropdown_name == "Double Clef":
                self.DOUBLE_CLE_OPTIONS = new_options
            elif dropdown_name == "Société Propriétaire":
                self.SOCIETE_PROPRIETAIRE_OPTIONS = new_options
            edit_dropdown_dialog.open = False
            self.page.update()

        edit_dropdown_dialog = ft.AlertDialog(
            title=ft.Text(f"Modifier les options de {dropdown_name}"),
            content=ft.Container(
                content=options_field,
                padding=20,
                width=350
            ),
            actions=[
                ft.TextButton("Annuler", on_click=lambda _: setattr(edit_dropdown_dialog, 'open', False)),
                ft.TextButton("Sauvegarder", on_click=save_options),
            ],
        )

        self.page.dialog = edit_dropdown_dialog
        edit_dropdown_dialog.open = True
        self.page.update()

    def toggle_theme_mode(self, e):
        self.page.theme_mode = (
            ft.ThemeMode.LIGHT
            if self.page.theme_mode == ft.ThemeMode.DARK
            else ft.ThemeMode.DARK
        )
        self.page.update()

    def show_export_dialog(self, e):
        """
        Affiche une boîte de dialogue permettant d'exporter les données
        au format Excel, CSV ou PDF.
        """

        def export_data(e):
            """
            Gère l'exportation des données selon le format choisi.
            """
            try:
                format = export_format_field.value
                if format not in ["Excel", "CSV", "PDF"]:
                    raise ValueError("Format d'exportation non pris en charge.")

                # Utiliser tkinter pour la sélection du fichier
                root = tk.Tk()
                root.withdraw()  # Cacher la fenêtre principale de tkinter
                root.attributes('-topmost', True)  # Mettre la fenêtre au premier plan
                file_path = filedialog.asksaveasfilename(
                    defaultextension=f".{format.lower()}",
                    filetypes=[(f"{format} files", f"*.{format.lower()}")]
                )
                root.attributes('-topmost', False)

                if not file_path:
                    return

                # Récupération des données
                vehicles = self.vehicle_repository.fetch_all_vehicles()
                df = pd.DataFrame([vars(vehicle) for vehicle in vehicles])

                # Export en fonction du format
                if format == "Excel":
                    df.to_excel(file_path, index=False)
                elif format == "CSV":
                    df.to_csv(file_path, index=False)
                elif format == "PDF":
                    self.export_to_pdf(df, file_path)

                # Afficher le message de confirmation
                self._show_export_complete_dialog(format, file_path)
            except Exception as error:
                # Afficher un message en cas d'erreur
                self._show_error_dialog(str(error))
            finally:
                export_dialog.open = False
                self.page.update()

        # Création du champ de sélection de format
        export_format_field = ft.Dropdown(
            label="Format",
            options=[
                ft.dropdown.Option("Excel"),
                ft.dropdown.Option("CSV"),
                ft.dropdown.Option("PDF")
            ],
            value="Excel"
        )

        # Boîte de dialogue principale
        export_dialog = ft.AlertDialog(
            title=ft.Text("Exporter les données"),
            content=ft.Column([export_format_field], spacing=10),
            actions=[
                ft.TextButton("Annuler", on_click=lambda _: setattr(export_dialog, 'open', False)),
                ft.TextButton("Exporter", on_click=export_data),
            ],
        )

        # Affichage de la boîte de dialogue
        self.page.dialog = export_dialog
        export_dialog.open = True
        self.page.update()

        # Mettre la fenêtre au premier plan
        self.page.window_to_front()

    def export_to_pdf(self, df: pd.DataFrame, file_path: str):
        """
        Exporte un DataFrame Pandas au format PDF, en s'assurant que toutes les colonnes sont incluses.
        """
        try:
            # Utiliser l'orientation paysage pour avoir plus d'espace horizontal
            doc = SimpleDocTemplate(file_path, pagesize=landscape(A4))
            elements = []

            # Préparation des données pour le tableau
            data = [df.columns.tolist()] + df.values.tolist()

            # Calculer la largeur disponible
            available_width = doc.width - inch  # Soustraire une marge

            # Calculer la largeur de chaque colonne
            col_widths = [available_width / len(df.columns)] * len(df.columns)

            # Création du tableau avec les largeurs de colonnes spécifiées
            table = Table(data, colWidths=col_widths, repeatRows=1)

            # Stylisation du tableau
            style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), pdf_colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), pdf_colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),  # Réduire la taille de la police pour les en-têtes
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), pdf_colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, pdf_colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),  # Réduire la taille de la police pour le contenu
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [pdf_colors.whitesmoke, pdf_colors.lightgrey])
            ])
            table.setStyle(style)
            elements.append(table)

            # Génération du PDF
            doc.build(elements)
        except Exception as e:
            raise RuntimeError(f"Erreur lors de l'exportation en PDF : {e}")

    def _show_export_complete_dialog(self, format: str, file_path: str):
        """
        Affiche une boîte de dialogue de confirmation après une exportation réussie.
        """

        def open_file():
            if os.path.exists(file_path):
                if os.name == 'nt':  # Windows
                    os.startfile(file_path)
                elif os.name == 'posix':  # macOS et Linux
                    subprocess.call(('open', file_path))
            export_complete_dialog.open = False
            self.page.update()

        export_complete_dialog = ft.AlertDialog(
            title=ft.Text("Exportation terminée"),
            content=ft.Text(
                f"Les données ont été exportées au format {format} sous le nom {os.path.basename(file_path)}."),
            actions=[
                ft.TextButton("Ouvrir le fichier", on_click=lambda _: open_file()),
                ft.TextButton("OK", on_click=lambda _: setattr(export_complete_dialog, 'open', False))
            ],
        )
        self.page.dialog = export_complete_dialog
        export_complete_dialog.open = True
        self.page.update()

    def _show_error_dialog(self, message: str):
        """
        Affiche une boîte de dialogue d'erreur.
        """
        error_dialog = ft.AlertDialog(
            title=ft.Text("Erreur"),
            content=ft.Text(f"Une erreur est survenue : {message}"),
            actions=[
                ft.TextButton("OK", on_click=lambda _: setattr(error_dialog, 'open', False))
            ],
        )
        self.page.dialog = error_dialog
        error_dialog.open = True
        self.page.update()

    def calculate_maintenance_count(self):
        vehicles = self.vehicle_repository.fetch_all_vehicles()
        today = datetime.today()
        maintenance_count = 0

        for vehicle in vehicles:
            try:
                derniere_revision_kms = int(vehicle.derniere_revision or 0)
                periodicite_revision_kms = int(vehicle.periodicite_revision or 0)
                releve_kms = int(vehicle.releve_kms or 0)

                prochaine_revision_kms = derniere_revision_kms + periodicite_revision_kms
                kms_difference = prochaine_revision_kms - releve_kms

                if isinstance(vehicle.date_derniere_revision, datetime):
                    derniere_revision_date = vehicle.date_derniere_revision
                else:
                    derniere_revision_date = datetime.strptime(vehicle.date_derniere_revision, '%d/%m/%Y')

                days_difference = (derniere_revision_date + timedelta(days=365)) - today
                days_remaining = days_difference.days

                if 0 <= kms_difference <= 1000 or 0 < days_remaining <= 45:
                    maintenance_count += 1
                elif kms_difference < 0 or days_remaining < 0:
                    maintenance_count += 1
            except (ValueError, AttributeError):
                continue

        return maintenance_count

    def calculate_ct_count(self):
        vehicles = self.vehicle_repository.fetch_all_vehicles()
        today = datetime.today()
        ct_count = 0

        for vehicle in vehicles:
            if vehicle.prochain_ct:
                try:
                    if isinstance(vehicle.prochain_ct, str):
                        prochain_ct_date = datetime.strptime(vehicle.prochain_ct, '%d/%m/%Y')
                    else:
                        prochain_ct_date = vehicle.prochain_ct
                except ValueError:
                    continue

                difference = (prochain_ct_date - today).days
                if difference <= 60:
                    ct_count += 1

        return ct_count

    def show_maintenance_alerts(self):
        maintenance_info = self.calculate_maintenance_info()
        alert_view = ft.ListView(spacing=10, padding=20, auto_scroll=True)

        if not maintenance_info:
            alert_view.controls.append(
                ft.Card(
                    content=ft.Container(
                        content=ft.Text(
                            "Aucun véhicule ne nécessite d'entretien.",
                            size=20,
                            color=colors.GREEN,
                            weight=ft.FontWeight.BOLD
                        ),
                        padding=20,
                        alignment=ft.alignment.center
                    )
                )
            )
        else:
            for info in maintenance_info:
                vehicle = info["vehicle"]
                prochaine_revision_kms = info["prochaine_revision_kms"]
                kms_difference = info["kms_difference"]
                days_remaining = info["days_remaining"]
                status_color = info["status_color"]

                alert_card = ft.Card(
                    content=ft.Container(
                        content=ft.Column([
                            ft.ListTile(
                                leading=ft.Icon(icons.DIRECTIONS_CAR, size=40, color=status_color),
                                height=60,
                                title=ft.Text(
                                    f"{vehicle.marque} {vehicle.vehicule} // {vehicle.immatriculation}",
                                    size=20,
                                    weight=ft.FontWeight.BOLD
                                ),
                                subtitle=ft.Column([
                                    ft.Text(f"Utilisateur: {vehicle.utilisateur}", size=12, weight=ft.FontWeight.BOLD),
                                    ft.Text(f"Site: {vehicle.site}", size=12, weight=ft.FontWeight.BOLD),
                                ]),
                            ),
                            ft.Container(
                                content=ft.Column([
                                    ft.Text(f"Plaque d'immatriculation: {vehicle.immatriculation}", size=14,
                                            weight=ft.FontWeight.BOLD),
                                    ft.Text(f"Date de la prochaine révision: {vehicle.date_derniere_revision}", size=14,
                                            weight=ft.FontWeight.BOLD),
                                    ft.Text(f"Kilométrage de la prochaine révision: {prochaine_revision_kms} km",
                                            size=14, weight=ft.FontWeight.BOLD),
                                    ft.Text(f"Kilomètres restants: {kms_difference} km", size=14,
                                            weight=ft.FontWeight.BOLD),
                                    ft.Text(f"Jours restants: {days_remaining} jours", size=14,
                                            weight=ft.FontWeight.BOLD),
                                ]),
                                padding=10,
                            ),
                        ]),
                        padding=5,
                    )
                )
                alert_view.controls.append(alert_card)

        self.main_content.content = alert_view
        self.page.update()

    def calculate_maintenance_info(self):
        vehicles = self.vehicle_repository.fetch_all_vehicles()
        today = datetime.today()
        maintenance_info = []

        for vehicle in vehicles:
            try:
                derniere_revision_kms = int(vehicle.derniere_revision or 0)
                periodicite_revision_kms = int(vehicle.periodicite_revision or 0)
                releve_kms = int(vehicle.releve_kms or 0)

                prochaine_revision_kms = derniere_revision_kms + periodicite_revision_kms
                kms_difference = prochaine_revision_kms - releve_kms

                if isinstance(vehicle.date_derniere_revision, datetime):
                    derniere_revision_date = vehicle.date_derniere_revision
                else:
                    derniere_revision_date = datetime.strptime(vehicle.date_derniere_revision, '%d/%m/%Y')

                days_difference = (derniere_revision_date + timedelta(days=365)) - today
                days_remaining = days_difference.days

                if 0 <= kms_difference <= 1000 or 0 < days_remaining <= 45:
                    status_color = colors.ORANGE
                elif kms_difference < 0 or days_remaining < 0:
                    status_color = colors.RED
                else:
                    continue

                maintenance_info.append({
                    "vehicle": vehicle,
                    "prochaine_revision_kms": prochaine_revision_kms,
                    "kms_difference": kms_difference,
                    "days_remaining": days_remaining,
                    "status_color": status_color,
                })
            except (ValueError, AttributeError):
                continue

        return maintenance_info

    def show_ct_alerts(self):
        ct_info = self.calculate_ct_info()
        alert_view = ft.ListView(spacing=10, padding=20, auto_scroll=True)

        if not ct_info:
            alert_view.controls.append(
                ft.Card(
                    content=ft.Container(
                        content=ft.Text(
                            "Aucun véhicule ne nécessite un contrôle technique.",
                            size=20,
                            color=colors.GREEN,
                            weight=ft.FontWeight.BOLD
                        ),
                        padding=20,
                        alignment=ft.alignment.center
                    )
                )
            )
        else:
            for vehicle, immatriculation, prochain_ct_date, status_message, icon_color in ct_info:
                alert_card = ft.Card(
                    content=ft.Container(
                        content=ft.Column([
                            ft.ListTile(
                                leading=ft.Icon(icons.DIRECTIONS_CAR, size=40, color=icon_color),
                                height=60,
                                title=ft.Text(
                                    f"{vehicle.marque} {vehicle.vehicule} // {immatriculation}",
                                    size=20,
                                    weight=ft.FontWeight.BOLD
                                ),
                                subtitle=ft.Column([
                                    ft.Text(f"Utilisateur: {vehicle.utilisateur}", size=10,
                                            weight=ft.FontWeight.BOLD),
                                    ft.Text(f"Site: {vehicle.site}", size=10, weight=ft.FontWeight.BOLD),
                                ]),
                            ),
                            ft.Container(
                                content=ft.Column([
                                    ft.Text(f"Plaque d'immatriculation: {immatriculation}", size=14,
                                            weight=ft.FontWeight.BOLD),
                                    ft.Text(f"Date du prochain contrôle technique: {prochain_ct_date}", size=14,
                                            weight=ft.FontWeight.BOLD),
                                    ft.Text(status_message, size=14, weight=ft.FontWeight.BOLD),
                                ]),
                                padding=10,
                            ),
                        ]),
                        padding=5,
                    )
                )
                alert_view.controls.append(alert_card)
        self.main_content.content = alert_view
        self.page.update()

    def calculate_ct_info(self):
        vehicles = self.vehicle_repository.fetch_all_vehicles()
        today = datetime.today()
        ct_info = []

        for vehicle in vehicles:
            if vehicle.prochain_ct:
                try:
                    if isinstance(vehicle.prochain_ct, str):
                        prochain_ct_date = datetime.strptime(vehicle.prochain_ct, '%d/%m/%Y')
                    else:
                        prochain_ct_date = vehicle.prochain_ct
                except ValueError:
                    continue

                difference = (prochain_ct_date - today).days
                days_remaining = abs(difference)
                icon_color = colors.ORANGE if 0 < difference <= 60 else colors.RED if difference < 0 else colors.GREEN

                if difference <= 60:
                    status_message = (
                        f"C.T valide pour encore {days_remaining} jours" if difference > 0 else
                        f"C.T périmé depuis {days_remaining} jours"
                    )
                    ct_info.append((vehicle, vehicle.immatriculation, prochain_ct_date.strftime('%d/%m/%Y'),
                                    status_message, icon_color))

        return ct_info

def main(page: ft.Page):
    vehicle_repository = VehicleRepository()
    VehicleManagementApp(page, vehicle_repository)

ft.app(target=main, view=ft.AppView.WEB_BROWSER)
